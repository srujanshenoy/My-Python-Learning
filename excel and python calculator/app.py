import openpyxl as xl

# def outcell_to_sum(cell,incell,outcell):
#     pass


workbook = xl.load_workbook("excelsheet.xlsx")
sheet = workbook["Sheet1"]

operator_cell = sheet.cell(2, 2)
operator_cell = operator_cell.value
#definitions
#cells
# define in_1
in_1 = sheet.cell(2, 1)
in_1 = in_1.value
in_1 = str(in_1)
in_1 = float(in_1)
# end

# define in_2
in_2 = sheet.cell(2, 3)
in_2 = in_2.value
in_2 = str(in_2)
in_2 = float(in_2)
# end

outcell = sheet.cell(2, 5)
#end


# functions
def add():
    sum_of_ins = in_1 + in_2
    outcell.value = sum_of_ins

def subtract():
    difference_of_ins = in_1 - in_2
    outcell.value = difference_of_ins

def multiply():
    product_of_ins = in_1 * in_2
    outcell.value = product_of_ins

def divide():
    if in_2 == 0:
        outcell.value = "Error: cannot divide by zero"
    else:
        quotient_of_ins = in_1 / in_2
        outcell.value = quotient_of_ins

def exponentiate():
    power_of_ins = in_1 ** in_2
    outcell.value = power_of_ins

#end

#end



# if operator_cell == "+":
#     sum_of_ins = in_1 + in_2
#     outcell.value = sum_of_ins


if operator_cell == "+":
    add()

# elif operator_cell == "-":
#     difference_of_ins = in_1 - in_2
#     outcell.value = difference_of_ins

elif operator_cell == "-":
    subtract()

# elif operator_cell == "x":
#     product_of_ins = in_1 * in_2
#     outcell.value = product_of_ins

elif operator_cell == "x":
    multiply()

# elif operator_cell == "/":
#     if in_2 == 0:
#         outcell.value = "Error: cannot divide by zero"
#     else:
#         quotient_of_ins = in_1 / in_2
#         outcell.value = quotient_of_ins

elif operator_cell == "/":
    divide()

# elif operator_cell == "^":
#     power_of_ins = in_1 ** in_2
#     outcell.value = power_of_ins

elif operator_cell == "^":
    exponentiate()
else:
    outcell.value = "Error: operator not supported."


# # elif in_1.__class__ == "str":
#     outcell.value = "kindly enter a different value for input 1"
# elif in_2.__class__ == "str":
#     outcell.value = "kindly enter a different value for input 2"
# elif in_1.__class__ == str &  in_2.__class__ == str:
#     outcell.value = "kindly enter a different value for input 1 and 2"

workbook.save("excelsheet.xlsx")
# print(outcell.value)#

